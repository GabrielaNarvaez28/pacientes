from django.forms import modelform_factory
from django.http import HttpResponse
from django.template import loader
from django.shortcuts import render, redirect, get_object_or_404
from openpyxl.workbook import Workbook
from rest_framework import viewsets, permissions

from pacientes.forms import PacienteFormulario
from pacientes.models import Paciente, Especialidad, Doctor
from pacientes.serializers import PacienteSerializer, EspecialidadSerializer, DoctorSerializer


# Create your views here.

#PacienteFormulario = modelform_factory(Paciente, exclude=['activo',])
def agregar_paciente(request):
    pagina = loader.get_template('pacientes/agregar.html')
    if request.method == 'GET':
        formulario = PacienteFormulario
    elif request.method == 'POST':
        formulario = PacienteFormulario(request.POST)
        if formulario.is_valid():
            formulario.save()
            return redirect('inicio')
    datos = {'formulario': formulario}
    return HttpResponse(pagina.render(datos,request))

def modificar_paciente(request, id):
    pagina = loader.get_template('pacientes/modificar.html')
    paciente = get_object_or_404(Paciente, pk=id)
    if request.method == 'GET':
        formulario = PacienteFormulario(instance=paciente)
    elif request.method == 'POST':
        formulario = PacienteFormulario(request.POST, instance=paciente)
        if formulario.is_valid():
            formulario.save()
            return redirect('inicio')
    datos = {'formulario': formulario}
    return HttpResponse(pagina.render(datos, request))

def ver_paciente(request, id):
    # paciente = Paciente.objects.get(pk=id)
    paciente = get_object_or_404(Paciente, pk=id)
    datos = {'paciente':paciente}
    # print(paciente)
    pagina = loader.get_template('pacientes/ver.html')
    return HttpResponse(pagina.render(datos, request))

def eliminar_paciente(request, id):
    paciente = get_object_or_404(Paciente, pk=id)
    if paciente:
        paciente.delete()
        return redirect('inicio')


def generar_reporte(request, *args, **kwargs):
    #Obtenemos todas las personas de nuestra base de datos
    pacientes = Paciente.objects.order_by('apellido', 'nombre')
    #Creamos el libro de trabajo
    wb = Workbook()
    #Definimos como nuestra hoja de trabajo, la hoja activa, por defecto la primera del libro
    ws = wb.active
    #En la celda B1 ponemos el texto 'REPORTE DE PERSONAS'
    ws['B1'] = 'REPORTE DE PACIENTES'
    #Juntamos las celdas desde la B1 hasta la E1, formando una sola celda
    ws.merge_cells('B1:E1')
    #Creamos los encabezados desde la celda B3 hasta la E3
    ws['B3'] = 'ID'
    ws['C3'] = 'NOMBRE'
    ws['D3'] = 'APELLIDO '
    ws['E3'] = 'CORREO'
    ws['F3'] = 'SEXO'
    cont=4
    #Recorremos el conjunto de personas y vamos escribiendo cada uno de los datos en las celdas
    for paciente in pacientes:
        ws.cell(row=cont,column=2).value = paciente.id
        ws.cell(row=cont,column=3).value = paciente.nombre
        ws.cell(row=cont,column=4).value = paciente.apellido
        ws.cell(row=cont,column=5).value = paciente.email
        ws.cell(row=cont,column=6).value = paciente.sexo
        cont = cont + 1
           #Establecemos el nombre del archivo
        nombre_archivo ="ReportePacientessExcel.xlsx"
           #Definimos que el tipo de respuesta a devolver es un archivo de microsoft excel
        response = HttpResponse(content_type="application/ms-excel")
        contenido = "attachment; filename={0}".format(nombre_archivo)
    response["Content-Disposition"] = contenido
    wb.save(response)
    return response



class PacienteViewSet(viewsets.ModelViewSet):
    """
    API endpoint that allows users to be viewed or edited.
    """
    queryset = Paciente.objects.all().order_by('-apellido')
    serializer_class = PacienteSerializer
    permission_classes = [permissions.IsAuthenticated]



class EspecialidadViewSet(viewsets.ModelViewSet):
    """
    API endpoint that allows users to be viewed or edited.
    """
    queryset = Especialidad.objects.all()
    serializer_class = EspecialidadSerializer
    permission_classes = [permissions.IsAuthenticated]


class DoctorViewSet(viewsets.ModelViewSet):
    """
    API endpoint that allows users to be viewed or edited.
    """
    queryset = Doctor.objects.all()
    serializer_class = DoctorSerializer
    permission_classes = [permissions.IsAuthenticated]









